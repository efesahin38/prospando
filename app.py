from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import calendar
import psycopg2
from psycopg2.extras import RealDictCursor


load_dotenv()

app = Flask(__name__)
app.secret_key = 'prospando-admin-2025'
CORS(app)

def get_conn():
    try:
        DATABASE_URL = os.getenv('DATABASE_URL')
        if not DATABASE_URL:
            raise ValueError("DATABASE_URL environment variable not found")
        conn = psycopg2.connect(
            DATABASE_URL,
            sslmode="require",
            connect_timeout=10
        )
        print("✅ Supabase bağlantısı başarılı!")
        return conn
    except Exception as e:
        print(f"❌ Bağlantı hatası: {str(e)}")
        raise

def calculate_hours(start_time, end_time):
    try:
        if not start_time or not end_time:
            return 0
        start = datetime.strptime(str(start_time)[:5], "%H:%M")
        end = datetime.strptime(str(end_time)[:5], "%H:%M")
        if end < start:
            end += timedelta(days=1)
        minutes = int((end - start).total_seconds() / 60)
        return round(minutes / 60, 2)
    except:
        return 0

def get_month_days(year, month):
    return calendar.monthrange(year, month)[1]

# ==================== EXCEL RAPOR ====================
@app.route('/api/export-all-employees-report', methods=['GET'])
def export_all_employees_report():
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Tüm çalışanları al
        cur.execute("SELECT id, name FROM employees ORDER BY name")
        employees = cur.fetchall()
        
        # Excel çalışma kitabı oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Mitarbeiterbericht"
        
        # Başlık stileri
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık
        ws['A1'] = "ALLE MITARBEITER - GESAMTE ARBEITSSTUNDEN"
        ws['A1'].font = Font(bold=True, size=14, color="7C3AED")
        ws.merge_cells('A1:B1')
        
        # Tablo başlıkları
        ws['A3'] = "Ad Soyad"
        ws['B3'] = "Toplam Saat"
        
        for col in ['A', 'B']:
            cell = ws[f'{col}3']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Kolon genişlikleri
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        # Veriler
        row_num = 4
        total_all_hours = 0
        
        for emp in employees:
            emp_id = emp['id']
            emp_name = emp['name']
            
            # Toplam çalışma saatini hesapla
            cur.execute("""
                SELECT start_time, end_time FROM attendance
                WHERE employee_id = %s AND end_time IS NOT NULL
            """, (emp_id,))
            
            total_hours = sum(calculate_hours(r['start_time'], r['end_time']) for r in cur.fetchall())
            total_all_hours += total_hours
            
            # Satıra ekle
            ws.cell(row=row_num, column=1).value = emp_name
            ws.cell(row=row_num, column=2).value = round(total_hours, 2)
            
            # Stil uygula
            for col in range(1, 3):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col == 2:
                    cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
        
        # Toplam satırı
        ws.cell(row=row_num, column=1).value = "GENEL TOPLAM"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).value = round(total_all_hours, 2)
        ws.cell(row=row_num, column=2).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row_num, column=2).fill = PatternFill(start_color="4ADE80", end_color="4ADE80", fill_type="solid")
        
        for col in range(1, 3):
            ws.cell(row=row_num, column=col).border = border
        
        cur.close()
        conn.close()
        
        # Excel dosyasını BytesIO'ya yaz
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Dosya adı
        filename = f"Alle Mitarbeiterberichte.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Excel export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== EMPLOYEES ====================

# ==================== EMPLOYEES ====================
@app.route('/api/employees', methods=['GET'])
def get_employees():
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT id, name FROM employees ORDER BY name")
        employees = []
        for emp in cur.fetchall():
            emp_id = emp['id']
            name = emp['name']
            cur.execute("SELECT email FROM users WHERE LOWER(name) = LOWER(%s) LIMIT 1", (name,))
            email_row = cur.fetchone()
            email = email_row['email'] if email_row else ''

            cur.execute("""
                SELECT start_time, end_time FROM attendance
                WHERE employee_id = %s AND end_time IS NOT NULL
            """, (emp_id,))
            total_hours = sum(calculate_hours(r['start_time'], r['end_time']) for r in cur.fetchall())

            employees.append({
                'id': emp_id,
                'name': name,
                'email': email,
                'total_hours': round(total_hours, 2)
            })
        cur.close()
        conn.close()
        return jsonify({'success': True, 'data': employees}), 200
    except Exception as e:
        print(f"❌ Employees error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employees', methods=['POST'])
def add_employee():
    try:
        data = request.json
        if not data or 'name' not in data:
            return jsonify({'success': False, 'error': 'Vorname Nachname ist Pflicht'}), 400

        name = data['name'].strip()

        if not name:
            return jsonify({'success': False, 'error': 'Die Felder für Vor- und Nachname dürfen nicht leer gelassen werden.'}), 400

        email = data.get('email')
        if email:
            email = email.strip()
            if not email:
                email = None
        else:
            email = None

        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT id FROM employees WHERE LOWER(name) = LOWER(%s)", (name,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Es gibt bereits einen Mitarbeiter mit diesem Namen.'}), 400

        cur.execute("SELECT MAX(id) AS max_id FROM employees")
        row = cur.fetchone()
        new_id = (row['max_id'] or 0) + 1

        cur.execute("INSERT INTO employees (id, name) VALUES (%s, %s)", (new_id, name))

        if email:
            cur.execute("""
                INSERT INTO users (name, email) 
                VALUES (%s, %s) 
                ON CONFLICT (email) DO UPDATE SET name = EXCLUDED.name
            """, (name, email))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'id': new_id,
                'name': name,
                'email': email or '',
                'total_hours': 0
            }
        }), 201

    except Exception as e:
        print(f"❌ Çalışan ekleme hatası: {str(e)}")
        return jsonify({'success': False, 'error': 'Serverfehler'}), 500

@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
def delete_employee(emp_id):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM attendance WHERE employee_id = %s", (emp_id,))
        cur.execute("DELETE FROM employees WHERE id = %s", (emp_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True}), 200
    except Exception as e:
        print(f"❌ Delete error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/<int:emp_id>', methods=['GET'])
def get_attendance(emp_id):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT a.date, a.start_time, a.end_time, a.location
            FROM attendance a
            WHERE a.employee_id = %s
            ORDER BY a.date DESC, a.start_time DESC
            LIMIT 100
        """, (emp_id,))
        rows = cur.fetchall()
        attendances = []
        for row in rows:
            hours = calculate_hours(row['start_time'], row['end_time']) if row['end_time'] else None
            attendances.append({
                'date': row['date'],
                'start_time': str(row['start_time'])[:5] if row['start_time'] else None,
                'end_time': str(row['end_time'])[:5] if row['end_time'] else None,
                'location': row['location'] or 'Bilinmiyor',
                'hours': hours
            })
        cur.close()
        conn.close()
        return jsonify({'success': True, 'data': attendances}), 200
    except Exception as e:
        print(f"❌ Attendance error: {str(e)}")
        return jsonify({'success': True, 'data': []}), 200

@app.route('/api/monthly-hours/<int:emp_id>', methods=['GET'])
def get_monthly_hours(emp_id):
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        first_day = f"{year}-{month:02d}-01"
        last_day = f"{year}-{month:02d}-{get_month_days(year, month)}"

        cur.execute("""
            SELECT date, start_time, end_time, location
            FROM attendance
            WHERE employee_id = %s AND date BETWEEN %s AND %s
            ORDER BY date DESC
        """, (emp_id, first_day, last_day))
        rows = cur.fetchall()

        total_hours = 0
        attendances = []
        unique_dates = set()

        for row in rows:
            hours = calculate_hours(row['start_time'], row['end_time']) if row['end_time'] else 0
            total_hours += hours
            unique_dates.add(row['date'])
            attendances.append({
                'date': row['date'],
                'start_time': str(row['start_time'])[:5] if row['start_time'] else None,
                'end_time': str(row['end_time'])[:5] if row['end_time'] else None,
                'location': row['location'] or 'unbekannt',
                'hours': round(hours, 2) if hours else None
            })

        cur.close()
        conn.close()
        return jsonify({
            'success': True,
            'total_hours': round(total_hours, 2),
            'work_days': len(unique_dates),
            'data': attendances
        }), 200
    except Exception as e:
        print(f"❌ Monthly hours error: {str(e)}")
        return jsonify({'success': True, 'total_hours': 0, 'work_days': 0, 'data': []}), 200

@app.route('/api/today-attendance', methods=['GET'])
def get_today_attendance():
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT e.name AS employee_name, a.start_time, a.end_time, a.location
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.date = %s
            ORDER BY a.start_time DESC
        """, (today,))
        rows = cur.fetchall()
        data = []
        for row in rows:
            hours = calculate_hours(row['start_time'], row['end_time'])
            data.append({
                'employee_name': row['employee_name'],
                'start_time': str(row['start_time'])[:5] if row['start_time'] else None,
                'end_time': str(row['end_time'])[:5] if row['end_time'] else None,
                'location': row['location'] or 'unbekannt',
                'hours': round(hours, 2) if hours else 0
            })
        cur.close()
        conn.close()
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        print(f"❌ Today attendance error: {str(e)}")
        return jsonify({'success': True, 'data': []}), 200

@app.route('/api/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    try:
        conn = get_conn()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT COUNT(*) AS count FROM employees")
        total_employees = cur.fetchone()['count']

        today = datetime.now().strftime("%Y-%m-%d")
        cur.execute("""
            SELECT COUNT(DISTINCT employee_id) AS count
            FROM attendance WHERE date = %s
        """, (today,))
        checked_in_today = cur.fetchone()['count']

        now = datetime.now()
        first_day = now.strftime("%Y-%m-01")
        cur.execute("""
            SELECT start_time, end_time
            FROM attendance
            WHERE date >= %s AND end_time IS NOT NULL
        """, (first_day,))
        total_monthly_hours = sum(
            calculate_hours(r['start_time'], r['end_time']) for r in cur.fetchall()
        )

        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'total_employees': total_employees,
            'checked_in_today': checked_in_today,
            'absent_today': total_employees - checked_in_today,
            'total_monthly_hours': round(total_monthly_hours, 2)
        }), 200
    except Exception as e:
        print(f"❌ Dashboard stats error: {str(e)}")
        return jsonify({
            'success': True,
            'total_employees': 0,
            'checked_in_today': 0,
            'absent_today': 0,
            'total_monthly_hours': 0
        }), 200

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'healthy', 'database': 'supabase'}), 200

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('DEBUG', 'True') == 'True'
    print("🚀 PROSPANDO Admin Backend başlatılıyor...")
    app.run(host='0.0.0.0', port=port, debug=debug)



